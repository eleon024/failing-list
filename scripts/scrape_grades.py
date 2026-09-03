#!/usr/bin/env python3
"""
scrape_grades.py — pull grades out of a folder of "Individual Student Report" PDFs
and spit out a failing-list as XLSX + CSV + an HTML dashboard.

Usage:
    python scrape_grades.py ./reports
    python scrape_grades.py ./reports -o ./out --threshold 79 --term P1
    python scrape_grades.py ./reports --all-terms --recursive

Requires: pdfplumber, pandas, openpyxl
    pip install pdfplumber pandas openpyxl
"""

import argparse
import html
import re
import sys
from dataclasses import dataclass, asdict
from pathlib import Path

import pdfplumber
import pandas as pd

# ---------------------------------------------------------------- patterns ---

# "Individual Student Report Alcocer Rivera, Jesus"
RE_STUDENT = re.compile(r"^Individual Student Report\s+(.+?)\s*$")

# "Class: 8(B) US HIST RECN A (Semester 1) Teacher: Curry"
RE_CLASS = re.compile(
    r"^Class:\s*(?P<body>.+?)\s*\(\s*Semester\s*(?P<sem>\d+)\s*\)\s*Teacher:\s*(?P<teacher>.+?)\s*$"
)

# leading period token like "8(B)" or "1(A)" or just "4"
RE_PERIOD = re.compile(r"^(?P<period>\d{1,2}\([A-Z]\)|\d{1,2})\s+(?P<name>.+)$")

# "P1 87 87% 1 0 1 0 0"  /  "P1 85 85%"  /  "P2 I 0%"
RE_TERM = re.compile(
    r"^(?P<term>[A-Z]{1,3}\d{0,2})\s+(?P<grade>-?\d+(?:\.\d+)?|[A-Z]{1,3})\s+"
    r"(?P<pct>-?\d+(?:\.\d+)?)%(?P<rest>.*)$"
)

# footer: "09/02/26 1"
RE_FOOTER = re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}\s+\d+\s*$")

SKIP_LINES = {"final grade", "semester grade", "current grade"}
HEADER_TOKENS = ("Rpt. Term", "Rpt Term")

ATT_COLS = ["absent", "tardy", "missing", "late", "incomplete"]


# ------------------------------------------------------------------ model ---

@dataclass
class Record:
    student: str
    period: str
    class_name: str
    semester: str
    teacher: str
    term: str
    grade: float | None
    grade_raw: str
    percent: float | None
    absent: int | None
    tardy: int | None
    missing: int | None
    late: int | None
    incomplete: int | None
    source_file: str
    page: int


# ----------------------------------------------------------------- parsing ---

def clean_lines(text: str):
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line:
            continue
        if RE_FOOTER.match(line):
            continue
        if line.lower() in SKIP_LINES:
            continue
        if line.startswith(HEADER_TOKENS):
            continue
        yield line


def split_period(body: str):
    """'8(B) US HIST RECN A' -> ('8(B)', 'US HIST RECN A')"""
    m = RE_PERIOD.match(body)
    if m:
        return m.group("period"), m.group("name").strip()
    return "", body.strip()


def to_num(s, cast=float):
    try:
        return cast(s)
    except (TypeError, ValueError):
        return None


def parse_pdf(path: Path, carry_student: str | None = None):
    """
    Yield Records. A student's report can span multiple pages; only the first
    page carries the name header, so we carry the current student forward
    across pages (and, if you dump one student per file, across files too).
    """
    records = []
    student = carry_student
    ctx = None  # current class context

    with pdfplumber.open(path) as pdf:
        for pageno, page in enumerate(pdf.pages, start=1):
            for line in clean_lines(page.extract_text()):

                m = RE_STUDENT.match(line)
                if m:
                    student = m.group(1).strip()
                    ctx = None
                    continue

                m = RE_CLASS.match(line)
                if m:
                    period, cname = split_period(m.group("body"))
                    ctx = {
                        "period": period,
                        "class_name": cname,
                        "semester": m.group("sem"),
                        "teacher": m.group("teacher").strip(),
                    }
                    continue

                m = RE_TERM.match(line)
                if m and ctx:
                    tail = m.group("rest").split()
                    att = dict(zip(ATT_COLS, (to_num(x, int) for x in tail)))
                    records.append(
                        Record(
                            student=student or "UNKNOWN",
                            term=m.group("term"),
                            grade=to_num(m.group("grade")),
                            grade_raw=m.group("grade"),
                            percent=to_num(m.group("pct")),
                            source_file=path.name,
                            page=pageno,
                            **ctx,
                            **{c: att.get(c) for c in ATT_COLS},
                        )
                    )

    return records, student


def scrape_dir(folder: Path, recursive: bool = False):
    pattern = "**/*.pdf" if recursive else "*.pdf"
    files = sorted(folder.glob(pattern))
    if not files:
        sys.exit(f"No PDFs found in {folder}")

    rows, carry = [], None
    for f in files:
        try:
            recs, carry = parse_pdf(f, carry_student=carry)
        except Exception as e:  # keep going; report at the end
            print(f"  !! {f.name}: {e}", file=sys.stderr)
            continue
        if not recs:
            print(f"  ?? {f.name}: no grade rows matched", file=sys.stderr)
        rows.extend(recs)
        print(f"  .. {f.name}: {len(recs)} rows")

    return pd.DataFrame([asdict(r) for r in rows])


# ------------------------------------------------------------------ output ---

def pick_term(df: pd.DataFrame, term: str | None, all_terms: bool):
    if all_terms or df.empty:
        return df
    if term:
        out = df[df["term"].str.upper() == term.upper()]
        if out.empty:
            print(f"  !! no rows for term {term}; falling back to latest",
                  file=sys.stderr)
        else:
            return out
    # latest = last term row listed per student+class
    return df.groupby(["student", "period", "class_name"], as_index=False).tail(1)


def write_excel(path: Path, failing, by_student, by_teacher, all_rows, threshold):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        failing.to_excel(xl, sheet_name="Failing", index=False)
        by_student.to_excel(xl, sheet_name="By Student", index=False)
        by_teacher.to_excel(xl, sheet_name="By Teacher", index=False)
        all_rows.to_excel(xl, sheet_name="All Grades", index=False)

        head_fill = PatternFill("solid", fgColor="1F3864")
        head_font = Font(color="FFFFFF", bold=True)
        bad_fill = PatternFill("solid", fgColor="F8CBAD")

        for name in xl.book.sheetnames:
            ws = xl.book[name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = Alignment(vertical="center")
            # column widths
            for col in ws.columns:
                letter = get_column_letter(col[0].column)
                width = max((len(str(c.value)) for c in col if c.value is not None),
                            default=8)
                ws.column_dimensions[letter].width = min(max(width + 2, 10), 42)
            # shade failing grades
            headers = [c.value for c in ws[1]]
            if "grade" in headers:
                gi = headers.index("grade") + 1
                for row in ws.iter_rows(min_row=2, min_col=gi, max_col=gi):
                    for c in row:
                        if isinstance(c.value, (int, float)) and c.value <= threshold:
                            c.fill = bad_fill
                            c.font = Font(bold=True, color="9C0006")


DASH_CSS = """
:root{--ink:#1a1a1a;--mut:#6b6b6b;--line:#e3e3e3;--bad:#b3261e;--badbg:#fdecea;
--accent:#1f3864;}
*{box-sizing:border-box}
body{font:15px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
color:var(--ink);margin:0;padding:24px;background:#fafafa}
h1{font-size:22px;margin:0 0 4px}
.sub{color:var(--mut);font-size:13px;margin-bottom:20px}
.cards{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:24px}
.card{background:#fff;border:1px solid var(--line);border-radius:10px;
padding:14px 18px;min-width:130px;flex:1}
.card .n{font-size:26px;font-weight:700;color:var(--accent)}
.card .l{font-size:12px;color:var(--mut);text-transform:uppercase;letter-spacing:.5px}
h2{font-size:15px;text-transform:uppercase;letter-spacing:.6px;color:var(--mut);
margin:26px 0 8px}
table{width:100%;border-collapse:collapse;background:#fff;border:1px solid var(--line);
border-radius:10px;overflow:hidden;font-size:14px}
th{background:var(--accent);color:#fff;text-align:left;padding:9px 10px;
font-size:12px;text-transform:uppercase;letter-spacing:.4px;cursor:pointer}
td{padding:8px 10px;border-top:1px solid var(--line)}
tr:hover td{background:#f5f7fb}
.g{font-weight:700;color:var(--bad);background:var(--badbg);border-radius:5px;
padding:1px 7px;display:inline-block}
input{width:100%;padding:9px 12px;border:1px solid var(--line);border-radius:8px;
margin-bottom:10px;font-size:14px}
@media(max-width:600px){body{padding:14px}td,th{padding:7px 6px;font-size:13px}}
"""

DASH_JS = """
function filt(){const q=document.getElementById('q').value.toLowerCase();
document.querySelectorAll('#fail tbody tr').forEach(r=>{
r.style.display=r.innerText.toLowerCase().includes(q)?'':'none';});}
document.querySelectorAll('#fail th').forEach((th,i)=>th.onclick=()=>{
const tb=th.closest('table').tBodies[0];
const rows=[...tb.rows];const asc=th.dataset.asc!=='1';th.dataset.asc=asc?'1':'0';
rows.sort((a,b)=>{const x=a.cells[i].innerText,y=b.cells[i].innerText;
const nx=parseFloat(x),ny=parseFloat(y);
return(!isNaN(nx)&&!isNaN(ny)?nx-ny:x.localeCompare(y))*(asc?1:-1);});
rows.forEach(r=>tb.appendChild(r));});
"""


def write_dashboard(path: Path, failing, by_student, by_teacher, threshold, n_students):
    def td(v):
        return html.escape("" if pd.isna(v) else str(v))

    rows = "\n".join(
        "<tr>"
        f"<td>{td(r.student)}</td>"
        f"<td>{td(r.class_name)}</td>"
        f"<td>{td(r.period)}</td>"
        f"<td>{td(r.teacher)}</td>"
        f"<td><span class='g'>{td(r.grade_raw)}</span></td>"
        "</tr>"
        for r in failing.itertuples()
    ) or "<tr><td colspan=5>Nobody below the line. 🎉</td></tr>"

    def mini(df, a, b):
        return "\n".join(
            f"<tr><td>{td(getattr(r, a))}</td><td>{td(getattr(r, b))}</td></tr>"
            for r in df.itertuples()
        ) or "<tr><td colspan=2>—</td></tr>"

    doc = f"""<!doctype html><html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Failing Report</title><style>{DASH_CSS}</style></head><body>
<h1>Failing Report</h1>
<div class="sub">Grades at or below {threshold} &middot; generated {pd.Timestamp.now():%b %d, %Y %I:%M %p}</div>
<div class="cards">
<div class="card"><div class="n">{len(failing)}</div><div class="l">Failing grades</div></div>
<div class="card"><div class="n">{failing['student'].nunique()}</div><div class="l">Students affected</div></div>
<div class="card"><div class="n">{n_students}</div><div class="l">Students scanned</div></div>
<div class="card"><div class="n">{failing['teacher'].nunique()}</div><div class="l">Teachers</div></div>
</div>
<input id="q" placeholder="Filter by student, class, or teacher…" oninput="filt()">
<table id="fail"><thead><tr><th>Student</th><th>Class</th><th>Per</th>
<th>Teacher</th><th>Grade</th></tr></thead><tbody>{rows}</tbody></table>
<h2>Most classes failing</h2>
<table><thead><tr><th>Student</th><th>Count</th></tr></thead>
<tbody>{mini(by_student.head(15), 'student', 'failing_classes')}</tbody></table>
<h2>By teacher</h2>
<table><thead><tr><th>Teacher</th><th>Count</th></tr></thead>
<tbody>{mini(by_teacher.head(20), 'teacher', 'failing_students')}</tbody></table>
<script>{DASH_JS}</script></body></html>"""
    path.write_text(doc, encoding="utf-8")


# -------------------------------------------------------------------- main ---

def main():
    ap = argparse.ArgumentParser(description="Scrape student report PDFs for failing grades.")
    ap.add_argument("folder", type=Path, help="directory of PDFs")
    ap.add_argument("-o", "--out", type=Path, default=Path("."), help="output directory")
    ap.add_argument("-t", "--threshold", type=float, default=79,
                    help="grades <= this count as failing (default 79)")
    ap.add_argument("--term", help="only this reporting term, e.g. P1")
    ap.add_argument("--all-terms", action="store_true",
                    help="keep every term row instead of just the latest")
    ap.add_argument("-r", "--recursive", action="store_true")
    args = ap.parse_args()

    print(f"Scanning {args.folder}…")
    df = scrape_dir(args.folder, args.recursive)
    if df.empty:
        sys.exit("Parsed 0 grade rows — check that the PDFs have selectable text.")

    n_students = df["student"].nunique()
    scoped = pick_term(df, args.term, args.all_terms)

    failing = (
        scoped[scoped["grade"].notna() & (scoped["grade"] <= args.threshold)]
        .sort_values(["student", "grade"])
        .reset_index(drop=True)
    )

    by_student = (
        failing.groupby("student")
        .agg(failing_classes=("class_name", "count"),
             lowest_grade=("grade", "min"),
             classes=("class_name", lambda s: ", ".join(s)))
        .sort_values(["failing_classes", "lowest_grade"], ascending=[False, True])
        .reset_index()
    )

    by_teacher = (
        failing.groupby(["teacher", "class_name"])
        .agg(failing_students=("student", "count"), avg_grade=("grade", "mean"))
        .round(1)
        .sort_values("failing_students", ascending=False)
        .reset_index()
    )

    args.out.mkdir(parents=True, exist_ok=True)
    xlsx = args.out / "failing_report.xlsx"
    csv = args.out / "failing_report.csv"
    dash = args.out / "failing_dashboard.html"

    cols = ["student", "period", "class_name", "teacher", "term", "grade",
            "grade_raw", "percent", "absent", "tardy", "missing", "source_file"]
    failing[cols].to_csv(csv, index=False)
    write_excel(xlsx, failing[cols], by_student, by_teacher, df, args.threshold)
    write_dashboard(dash, failing, by_student, by_teacher, args.threshold, n_students)

    print(f"\n{len(df)} grade rows · {n_students} students")
    print(f"{len(failing)} failing grades across {failing['student'].nunique()} students")
    print(f"\n  {xlsx}\n  {csv}\n  {dash}")


if __name__ == "__main__":
    main()
